"""
Qdrant-backed Feedback Store for Sourcing Intelligence.
Stores candidate feedback as vectors for semantic similarity search.
Uses FastEmbed (ONNX) for automatic text embedding — no GPU required.
"""

import uuid
import io
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from qdrant_client import QdrantClient, models

# ── Constants ─────────────────────────────────────────────────────────────────
QDRANT_PATH = Path("data/qdrant_feedback")
COLLECTION_NAME = "candidate_feedback"

# Column name mappings from the real Excel sheet
COL_ROLE = "HSBC Role Mapping"
COL_CANDIDATE = "Resource Name"
COL_STATUS = "Status"
COL_FEEDBACK = "Client Feedback"
COL_SKILLS = "Skills / Domain"
COL_SUMMARY = "Profile Summary"
COL_EXPERIENCE = "Total / Relevant Experience"

# ── Qdrant Client (module-level singleton) ────────────────────────────────────
_qdrant_client: QdrantClient | None = None


def _get_client() -> QdrantClient:
    global _qdrant_client
    if _qdrant_client is None:
        raise RuntimeError("Feedback store not initialised. Call init_feedback_store() first.")
    return _qdrant_client


def init_feedback_store():
    """Initialise Qdrant client with local disk persistence."""
    global _qdrant_client
    if _qdrant_client is not None:
        return
    QDRANT_PATH.mkdir(parents=True, exist_ok=True)
    _qdrant_client = QdrantClient(path=str(QDRANT_PATH))


# ── Excel Ingestion ───────────────────────────────────────────────────────────

def _derive_feedback_type(status: str) -> str:
    """Derive 'positive' or 'negative' from the Status column value."""
    if not status:
        return "unknown"
    lower = status.strip().lower()
    if "reject" in lower:
        return "negative"
    return "positive"


def ingest_feedback_excel(file_bytes: bytes) -> dict:
    """
    Parse an .xlsx file and store each feedback row as a Qdrant point.

    Returns:
        {
            "processed": int,
            "skipped": int,
            "errors": list[str],
            "roles_found": list[str]
        }
    """
    client = _get_client()
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    # Read header row and build column index
    headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {h: i for i, h in enumerate(headers)}

    # Validate required columns exist
    required = [COL_ROLE, COL_CANDIDATE, COL_STATUS, COL_FEEDBACK]
    missing = [c for c in required if c not in col_map]
    if missing:
        return {
            "processed": 0,
            "skipped": 0,
            "errors": [f"Missing required columns: {', '.join(missing)}"],
            "roles_found": [],
        }

    # Reset collection (drop collection) to ensure fresh named vector configuration on next add()
    try:
        client.delete_collection(COLLECTION_NAME)
    except Exception as e:
        errors.append(f"Warning: Failed to clear old collection: {str(e)}")

    processed = 0
    skipped = 0
    errors = []
    roles_seen = set()
    points = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            role = str(row[col_map[COL_ROLE]] or "").strip()
            candidate = str(row[col_map[COL_CANDIDATE]] or "").strip()
            status_raw = str(row[col_map[COL_STATUS]] or "").strip()
            feedback = str(row[col_map[COL_FEEDBACK]] or "").strip()

            # Skip rows without both status and feedback
            if not status_raw or not feedback or status_raw == "None" or feedback == "None":
                skipped += 1
                continue

            if not role:
                skipped += 1
                continue

            feedback_type = _derive_feedback_type(status_raw)
            roles_seen.add(role)

            # Optional columns
            skills = str(row[col_map.get(COL_SKILLS, -1)] or "").strip() if COL_SKILLS in col_map else ""
            summary = str(row[col_map.get(COL_SUMMARY, -1)] or "").strip() if COL_SUMMARY in col_map else ""
            experience = str(row[col_map.get(COL_EXPERIENCE, -1)] or "").strip() if COL_EXPERIENCE in col_map else ""

            point_id = str(uuid.uuid4())
            payload = {
                "role": role,
                "candidate_name": candidate,
                "feedback_type": feedback_type,
                "feedback_text": feedback,
                "status_raw": status_raw,
                "skills_domain": skills,
                "profile_summary": summary,
                "experience": experience,
                "stored_at": datetime.utcnow().isoformat(),
            }

            points.append({
                "id": point_id,
                "text": feedback,  # This text will be embedded
                "payload": payload,
            })

            processed += 1

        except Exception as e:
            errors.append(f"Row error: {str(e)}")

    wb.close()

    # Batch upsert with auto-embedding via FastEmbed
    if points:
        from qdrant_client.models import PointStruct
        # Use the add() method with documents for auto-embedding
        client.add(
            collection_name=COLLECTION_NAME,
            documents=[p["text"] for p in points],
            metadata=[p["payload"] for p in points],
            ids=[p["id"] for p in points],
        )

    return {
        "processed": processed,
        "skipped": skipped,
        "errors": errors,
        "roles_found": sorted(roles_seen),
    }


# ── Query Rejection Trends ───────────────────────────────────────────────────

def query_rejection_trends(role: str, limit: int = 50) -> list[dict]:
    """
    Search Qdrant for negative feedback entries matching the given role.
    Uses strict payload filtering so roles don't mix.

    Returns list of payload dicts for matching points.
    """
    client = _get_client()

    # Strict filter: exact role match AND negative feedback only
    role_filter = models.Filter(
        must=[
            models.FieldCondition(
                key="role",
                match=models.MatchValue(value=role),
            ),
            models.FieldCondition(
                key="feedback_type",
                match=models.MatchValue(value="negative"),
            ),
        ]
    )

    # Scroll through all matching points (no vector query needed — we want ALL negative feedback for this role)
    try:
        results, _ = client.scroll(
            collection_name=COLLECTION_NAME,
            scroll_filter=role_filter,
            limit=limit,
            with_payload=True,
        )
    except Exception:
        # Collection does not exist yet (no files uploaded)
        return []

    return [point.payload for point in results]


# ── Feedback Stats ────────────────────────────────────────────────────────────

def get_feedback_stats() -> dict:
    """
    Returns feedback counts grouped by role and type.
    Format: { "roles": { "Gen AI BA": { "positive": 3, "negative": 5, "total": 8 }, ... }, "total": 20 }
    """
    client = _get_client()

    # Scroll all points to aggregate stats
    try:
        all_points, _ = client.scroll(
            collection_name=COLLECTION_NAME,
            limit=1000,
            with_payload=True,
        )
    except Exception:
        # Collection does not exist yet
        return {"roles": {}, "total": 0}

    roles = {}
    for point in all_points:
        role = point.payload.get("role", "Unknown")
        ftype = point.payload.get("feedback_type", "unknown")

        if role not in roles:
            roles[role] = {"positive": 0, "negative": 0, "total": 0}

        if ftype == "positive":
            roles[role]["positive"] += 1
        elif ftype == "negative":
            roles[role]["negative"] += 1
        roles[role]["total"] += 1

    return {
        "roles": roles,
        "total": sum(r["total"] for r in roles.values()),
    }


def clear_all_feedback() -> bool:
    """Delete the candidate_feedback collection completely from disk."""
    client = _get_client()
    try:
        client.delete_collection(COLLECTION_NAME)
        return True
    except Exception:
        return False
